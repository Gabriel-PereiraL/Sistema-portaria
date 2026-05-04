# =============================================================================
# ui.py — Interface gráfica completa (PyQt6)
# =============================================================================

import csv
import os
import sys
from datetime import datetime

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QStackedWidget,
    QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QComboBox,
    QTableWidget, QTableWidgetItem, QMessageBox,
    QHeaderView, QFrame, QScrollArea,
    QDialog, QGridLayout, QFileDialog,
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal
from PyQt6.QtGui import QColor, QPalette

from queries import (
    buscar_visitantes,
    buscar_funcionarios,
    buscar_visitantes_dentro,
    registrar_saida,
    buscar_historico_visitas,
    buscar_stats,
)
from logic import (
    LOGIN_USUARIO,
    LOGIN_SENHA,
    tentar_registrar_entrada,
    tentar_cadastrar_visitante,
)
from utils import formatar_tempo_permanencia


# =============================================================================
# HELPER DE EXPORTAÇÃO (usado por HistoricoPage e DentroPage)
# =============================================================================

def exportar_tabela_csv(parent: QWidget, tabela: QTableWidget, nome_sugerido: str):
    """
    Exporta os dados visíveis em `tabela` para CSV.
    Abre QFileDialog para escolha do destino.
    Não toca em queries nem em logic.
    """
    caminho, _ = QFileDialog.getSaveFileName(
        parent,
        "Salvar CSV",
        os.path.expanduser(f"~/{nome_sugerido}.csv"),
        "CSV (*.csv);;Todos os arquivos (*)",
    )
    if not caminho:
        return  # usuário cancelou

    headers = [
        tabela.horizontalHeaderItem(c).text()
        for c in range(tabela.columnCount())
    ]

    try:
        with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in range(tabela.rowCount()):
                writer.writerow(
                    tabela.item(row, col).text() if tabela.item(row, col) else ""
                    for col in range(tabela.columnCount())
                )
        QMessageBox.information(parent, "Exportado", f"Arquivo salvo em:\n{caminho}")
    except Exception as e:
        print(f"[ERRO] exportar_tabela_csv: {e}")
        QMessageBox.critical(parent, "Erro", f"Não foi possível salvar o arquivo.\n{e}")


def exportar_tabela_excel(parent: QWidget, tabela: QTableWidget, nome_sugerido: str):
    """
    Exporta os dados visíveis em `tabela` para .xlsx usando openpyxl.
    Só aparece no menu se openpyxl estiver instalado.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        QMessageBox.warning(
            parent, "openpyxl não instalado",
            "Instale com:  pip install openpyxl\nUsando exportação CSV como alternativa.",
        )
        exportar_tabela_csv(parent, tabela, nome_sugerido)
        return

    caminho, _ = QFileDialog.getSaveFileName(
        parent,
        "Salvar Excel",
        os.path.expanduser(f"~/{nome_sugerido}.xlsx"),
        "Excel (*.xlsx);;Todos os arquivos (*)",
    )
    if not caminho:
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = nome_sugerido[:31]  # limite do Excel

        # Estilo do cabeçalho
        header_fill = PatternFill("solid", fgColor="2D1B69")
        header_font = Font(color="FFFFFF", bold=True)

        headers = [
            tabela.horizontalHeaderItem(c).text()
            for c in range(tabela.columnCount())
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for row in range(tabela.rowCount()):
            for col in range(tabela.columnCount()):
                item = tabela.item(row, col)
                ws.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

        # Ajusta largura das colunas automaticamente
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(caminho)
        QMessageBox.information(parent, "Exportado", f"Arquivo salvo em:\n{caminho}")
    except Exception as e:
        print(f"[ERRO] exportar_tabela_excel: {e}")
        QMessageBox.critical(parent, "Erro", f"Não foi possível salvar o arquivo.\n{e}")


# =============================================================================
# STYLESHEET GLOBAL (QSS) — sem alterações
# =============================================================================

STYLESHEET = """
/* ── Reset & Base ── */


QMainWindow, QWidget {
    background-color: #0f0d1a;
    color: #e2e0f0;
    font-family: 'Segoe UI', 'Trebuchet MS', 'Gill Sans', sans-serif;
    font-size: 13px;
}

/* ── Sidebar ── */
#sidebar {
    background-color: #13101f;
    border-right: 1px solid #1e1b2e;
    min-width: 230px;
    max-width: 230px;
}
#sidebar_header {
    background-color: #13101f;
    border-bottom: 1px solid #1e1b2e;
}
#system_name {
    color: #a78bfa;
    font-size: 14px;
    font-weight: bold;
    letter-spacing: 0.5px;
}
#system_subtitle {
    color: #4b4580;
    font-size: 10px;
    letter-spacing: 1.5px;
    text-transform: uppercase;
}

/* ── Nav Items ── */
#nav_item {
    background: transparent;
    color: #6b5f8c;
    border: none;
    border-radius: 8px;
    padding: 10px 14px;
    text-align: left;
    font-size: 13px;
    font-weight: 500;
}
#nav_item:hover { background: #1e1b2e; color: #c4b5fd; }

#nav_item_active {
    background: #2d1b69;
    color: #ffffff;
    border: none;
    border-radius: 8px;
    padding: 10px 14px;
    text-align: left;
    font-size: 13px;
    font-weight: bold;
    border-left: 3px solid #7c3aed;
}

/* ── Área Principal ── */
#main_area { background-color: #0f0d1a; }

#topbar {
    background-color: #13101f;
    border-bottom: 1px solid #1e1b2e;
    min-height: 56px;
    max-height: 56px;
}
#page_title    { color: #e2e0f0; font-size: 18px; font-weight: bold; letter-spacing: -0.3px; }
#page_subtitle { color: #4b4580; font-size: 11px; }

#badge_timer {
    background: #1e1b2e; color: #6b5f8c; font-size: 11px;
    border: 1px solid #2a2550; border-radius: 20px; padding: 4px 12px;
}
#badge_online {
    background: #0a1f12; color: #4ade80; font-size: 11px;
    border: 1px solid #0d3320; border-radius: 20px; padding: 4px 12px;
}

/* ── Conteúdo ── */
#content_area { background-color: #0f0d1a; }

/* ── Cards ── */
#card {
    background: #13101f;
    border: 1px solid #1e1b2e;
    border-radius: 12px;
}
#stat_card {
    background: #13101f;
    border: 1px solid #1e1b2e;
    border-radius: 12px;
    min-height: 100px;
    max-height: 120px;
}
#stat_card:hover { border-color: #2d1b69; background: #160e30; }
#stat_value { font-size: 32px; font-weight: bold; color: #ffffff; letter-spacing: -1px; }
#stat_label { font-size: 12px; color: #6b5f8c; letter-spacing: 0.3px; }

/* ── Títulos de Seção ── */
#section_title {
    color: #9d8ec4; font-size: 11px; font-weight: bold;
    letter-spacing: 1.2px; text-transform: uppercase; padding-bottom: 2px;
}

/* ── Labels de Formulário ── */
#form_label {
    color: #6b5f8c; font-size: 11px; font-weight: bold;
    letter-spacing: 0.8px; text-transform: uppercase;
}

/* ── Inputs e ComboBoxes ── */
QLineEdit, QComboBox {
    background: #0f0d1a;
    border: 1.5px solid #1e1b2e;
    border-radius: 8px;
    padding: 10px 14px;
    color: #e2e0f0;
    font-size: 13px;
    selection-background-color: #7c3aed;
}
QLineEdit:focus, QComboBox:focus { border-color: #7c3aed; background: #110e1e; }
QLineEdit:hover, QComboBox:hover { border-color: #2d1b69; }
QLineEdit::placeholder { color: #3a3360; }

QComboBox::drop-down { border: none; width: 30px; }
QComboBox::down-arrow {
    image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid #6b5f8c;
    margin-right: 8px;
}
QComboBox QAbstractItemView {
    background: #13101f;
    border: 1.5px solid #2d1b69;
    border-radius: 8px;
    color: #e2e0f0;
    selection-background-color: #2d1b69;
    selection-color: #ffffff;
    padding: 4px;
    outline: none;
}

/* ── Botões ── */
QPushButton#btn_primary {
    background: #7c3aed; color: #ffffff;
    border: none; border-radius: 8px;
    padding: 10px 20px; font-weight: bold; font-size: 13px;
}
QPushButton#btn_primary:hover    { background: #8b5cf6; }
QPushButton#btn_primary:pressed  { background: #6d28d9; }
QPushButton#btn_primary:disabled { background: #1e1b2e; color: #3a3360; }

QPushButton#btn_secondary {
    background: transparent; color: #9d8ec4;
    border: 1.5px solid #1e1b2e; border-radius: 8px;
    padding: 9px 18px; font-size: 13px;
}
QPushButton#btn_secondary:hover  { background: #1e1b2e; border-color: #2d1b69; color: #c4b5fd; }
QPushButton#btn_secondary:pressed { background: #13101f; }

QPushButton#btn_danger {
    background: transparent; color: #f87171;
    border: 1.5px solid #2d1515; border-radius: 8px;
    padding: 9px 18px; font-size: 13px;
}
QPushButton#btn_danger:hover { background: #1a0a0a; border-color: #7c2020; color: #fca5a5; }

QPushButton#btn_success {
    background: #065f46; color: #4ade80;
    border: 1.5px solid #0a3d2a; border-radius: 8px;
    padding: 9px 18px; font-weight: bold; font-size: 13px;
}
QPushButton#btn_success:hover { background: #047857; border-color: #065f46; }

/* ── Tabelas ── */
QTableWidget {
    background: #0f0d1a;
    border: 1px solid #1e1b2e;
    gridline-color: #13101f;
    border-radius: 10px;
    alternate-background-color: #110e1e;
    outline: none;
    selection-background-color: #1e1b2e;
}
QTableWidget::item { padding: 10px 12px; color: #c4b5fd; border: none; }
QTableWidget::item:selected { background: #1e1b2e; color: #ffffff; }
QTableWidget::item:hover:!selected { background: #13101f; }

QHeaderView::section {
    background: #13101f; color: #6b5f8c;
    padding: 10px 12px; border: none;
    border-bottom: 2px solid #1e1b2e;
    font-weight: bold; font-size: 11px; letter-spacing: 0.8px;
    text-transform: uppercase;
}
QHeaderView::section:first { border-top-left-radius: 9px; }
QHeaderView::section:last  { border-top-right-radius: 9px; }
QTableCornerButton::section { background: #13101f; border: none; }

/* ── Scrollbars ── */
QScrollBar:vertical   { background: transparent; width: 6px; margin: 0; }
QScrollBar::handle:vertical { background: #1e1b2e; border-radius: 3px; min-height: 24px; }
QScrollBar::handle:vertical:hover { background: #2d1b69; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QScrollBar:horizontal { background: transparent; height: 6px; }
QScrollBar::handle:horizontal { background: #1e1b2e; border-radius: 3px; min-width: 24px; }
QScrollBar::handle:horizontal:hover { background: #2d1b69; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

/* ── QMessageBox ── */
QMessageBox { background-color: #13101f; color: #e2e0f0; }
QMessageBox QPushButton {
    min-width: 80px; background: #7c3aed; color: #fff;
    border: none; border-radius: 6px; padding: 7px 16px; font-weight: bold;
}
QMessageBox QPushButton:hover { background: #8b5cf6; }

/* ── Tooltip ── */
QToolTip {
    background: #1e1b2e; color: #e2e0f0;
    border: 1px solid #2d1b69; border-radius: 6px; padding: 5px 10px; font-size: 12px;
}

/* ── Divisor ── */
#divider { background: #1e1b2e; max-height: 1px; min-height: 1px; }

/* ── ScrollArea transparente ── */
QScrollArea { background: transparent; border: none; }
QScrollArea > QWidget > QWidget { background: transparent; }

/* ── Login ── */
#login_window   { background-color: #0f0d1a; }
#login_card     { background: #13101f; border: 1px solid #1e1b2e; border-radius: 16px; }
#login_title    { color: #a78bfa; font-size: 22px; font-weight: bold; letter-spacing: 0.5px; }
#login_subtitle { color: #4b4580; font-size: 11px; letter-spacing: 1.5px; }
#login_form_label {
    color: #6b5f8c; font-size: 11px; font-weight: bold;
    letter-spacing: 0.8px; text-transform: uppercase;
}
#login_error {
    color: #f87171; font-size: 12px;
    background: #1a0a0a; border: 1px solid #2d1515;
    border-radius: 8px; padding: 8px 14px;
}
QPushButton#btn_login {
    background: #7c3aed; color: #ffffff;
    border: none; border-radius: 8px;
    padding: 12px 20px; font-weight: bold; font-size: 14px; letter-spacing: 0.5px;
}
QPushButton#btn_login:hover   { background: #8b5cf6; }
QPushButton#btn_login:pressed { background: #6d28d9; }

/* ── Modal de Detalhes ── */
#detail_dialog { background-color: #13101f; }
#detail_header {
    background: #160e30;
    border-bottom: 1px solid #2d1b69;
    border-radius: 0px;
}
#detail_title    { color: #a78bfa; font-size: 15px; font-weight: bold; letter-spacing: 0.3px; }
#detail_subtitle { color: #4b4580; font-size: 11px; }
#detail_field_label {
    color: #6b5f8c; font-size: 10px; font-weight: bold;
    letter-spacing: 1px; text-transform: uppercase;
}
#detail_field_value {
    color: #e2e0f0; font-size: 13px;
    background: #0f0d1a; border: 1px solid #1e1b2e;
    border-radius: 8px; padding: 8px 12px;
}
#detail_field_value_desc {
    color: #c4b5fd; font-size: 13px;
    background: #0f0d1a; border: 1px solid #2d1b69;
    border-radius: 8px; padding: 8px 12px;
}
#detail_active_badge {
    color: #4ade80; font-size: 12px; font-weight: bold;
    background: #0a1f12; border: 1px solid #0d3320;
    border-radius: 20px; padding: 4px 14px;
}
#detail_done_badge {
    color: #6b5f8c; font-size: 12px;
    background: #13101f; border: 1px solid #1e1b2e;
    border-radius: 20px; padding: 4px 14px;
}

/* ── Contador de resultados de filtro ── */
#filtro_contador {
    color: #4b4580; font-size: 11px; background: transparent;
}
"""


# =============================================================================
# COMPONENTES REUTILIZÁVEIS — sem alterações
# =============================================================================

def make_divider():
    d = QFrame()
    d.setObjectName("divider")
    d.setFrameShape(QFrame.Shape.HLine)
    return d


def make_card(padding=20, spacing=14):
    card   = QWidget()
    card.setObjectName("card")
    layout = QVBoxLayout(card)
    layout.setContentsMargins(padding, padding, padding, padding)
    layout.setSpacing(spacing)
    return card, layout


def styled_input(placeholder=""):
    inp = QLineEdit()
    inp.setPlaceholderText(placeholder)
    inp.setMinimumHeight(40)
    return inp


def styled_combo():
    cb = QComboBox()
    cb.setMinimumHeight(40)
    return cb


def primary_button(text, height=40):
    btn = QPushButton(text)
    btn.setObjectName("btn_primary")
    btn.setMinimumHeight(height)
    return btn


def secondary_button(text, height=36):
    btn = QPushButton(text)
    btn.setObjectName("btn_secondary")
    btn.setMinimumHeight(height)
    return btn


def success_button(text, height=40):
    btn = QPushButton(text)
    btn.setObjectName("btn_success")
    btn.setMinimumHeight(height)
    return btn


def danger_button(text, height=36):
    btn = QPushButton(text)
    btn.setObjectName("btn_danger")
    btn.setMinimumHeight(height)
    return btn


def form_label(text):
    lbl = QLabel(text)
    lbl.setObjectName("form_label")
    return lbl


# =============================================================================
# MODAL DE DETALHES DA VISITA — sem alterações
# =============================================================================

class DetalhesVisitaDialog(QDialog):

    def __init__(self, dados: dict, parent=None):
        super().__init__(parent)
        self.setObjectName("detail_dialog")
        self.setWindowTitle("Detalhes da Visita")
        self.setStyleSheet(STYLESHEET)
        self.setMinimumWidth(540)
        self.setModal(True)
        self._build(dados)
        self._centralizar()

    def _build(self, d: dict):
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(0, 0, 0, 0)
        main_lay.setSpacing(0)

        header = QWidget()
        header.setObjectName("detail_header")
        h_lay = QHBoxLayout(header)
        h_lay.setContentsMargins(24, 18, 24, 18)
        h_lay.setSpacing(12)

        icon_lbl = QLabel("≡")
        icon_lbl.setStyleSheet("color: #7c3aed; font-size: 22px; background: transparent;")
        h_lay.addWidget(icon_lbl)

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        lbl_title = QLabel("Detalhes da Visita")
        lbl_title.setObjectName("detail_title")
        lbl_sub = QLabel(f"Registro #{d.get('id', '—')}")
        lbl_sub.setObjectName("detail_subtitle")
        title_col.addWidget(lbl_title)
        title_col.addWidget(lbl_sub)
        h_lay.addLayout(title_col)
        h_lay.addStretch()

        em = d.get("saida") is None or d.get("saida") == "—"
        badge = QLabel("🟢  Em andamento" if em else "✓  Concluído")
        badge.setObjectName("detail_active_badge" if em else "detail_done_badge")
        h_lay.addWidget(badge)
        main_lay.addWidget(header)

        body = QWidget()
        body.setObjectName("detail_dialog")
        body_lay = QVBoxLayout(body)
        body_lay.setContentsMargins(24, 20, 24, 24)
        body_lay.setSpacing(16)

        grid = QGridLayout()
        grid.setSpacing(12)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)

        def add_field(grid, row, col, label_txt, value_txt, obj="detail_field_value", colspan=1):
            col_lay = QVBoxLayout()
            col_lay.setSpacing(4)
            lbl = QLabel(label_txt)
            lbl.setObjectName("detail_field_label")
            val = QLabel(str(value_txt) if value_txt else "—")
            val.setObjectName(obj)
            val.setWordWrap(True)
            col_lay.addWidget(lbl)
            col_lay.addWidget(val)
            if colspan > 1:
                grid.addLayout(col_lay, row, col, 1, colspan)
            else:
                grid.addLayout(col_lay, row, col)

        add_field(grid, 0, 0, "VISITANTE",   d.get("visitante", "—"))
        add_field(grid, 0, 1, "FUNCIONÁRIO", d.get("funcionario", "—"))
        add_field(grid, 1, 0, "PORTEIRO",    d.get("porteiro", "—") or "—")
        add_field(grid, 1, 1, "SETOR",       d.get("setor", "—"))
        add_field(grid, 2, 0, "ENTRADA",     d.get("entrada", "—"))
        add_field(grid, 2, 1, "SAÍDA",       d.get("saida", "—") or "Em andamento")
        add_field(grid, 3, 0, "PERMANÊNCIA", d.get("permanencia", "—"))
        body_lay.addLayout(grid)

        body_lay.addWidget(make_divider())

        desc_lay = QVBoxLayout()
        desc_lay.setSpacing(6)
        lbl_desc = QLabel("DESCRIÇÃO COMPLETA")
        lbl_desc.setObjectName("detail_field_label")
        desc_lay.addWidget(lbl_desc)

        desc_text = QLabel(d.get("descricao") or "Nenhuma descrição informada.")
        desc_text.setObjectName("detail_field_value_desc")
        desc_text.setWordWrap(True)
        desc_text.setMinimumHeight(60)
        desc_text.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        desc_lay.addWidget(desc_text)
        body_lay.addLayout(desc_lay)

        body_lay.addSpacing(4)

        btn_fechar = primary_button("Fechar", height=40)
        btn_fechar.setMaximumWidth(120)
        btn_fechar.clicked.connect(self.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(btn_fechar)
        body_lay.addLayout(btn_row)

        main_lay.addWidget(body)

    def _centralizar(self):
        if self.parent():
            pg = self.parent().geometry()
            self.move(
                pg.x() + (pg.width()  - self.width())  // 2,
                pg.y() + (pg.height() - self.height()) // 2,
            )
        else:
            screen = QApplication.primaryScreen()
            if screen:
                geo = screen.availableGeometry()
                self.move(
                    (geo.width()  - self.width())  // 2,
                    (geo.height() - self.height()) // 2,
                )


# =============================================================================
# TELA DE LOGIN — sem alterações
# =============================================================================

class LoginWindow(QWidget):
    login_aceito = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setObjectName("login_window")
        self.setWindowTitle("Sistema de Portaria — Login")
        self.setFixedSize(440, 500)
        self._build()
        self._centralizar()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)

        center = QWidget()
        center.setObjectName("login_window")
        center_lay = QVBoxLayout(center)
        center_lay.setContentsMargins(40, 40, 40, 40)
        center_lay.addStretch()

        card = QWidget()
        card.setObjectName("login_card")
        card_lay = QVBoxLayout(card)
        card_lay.setContentsMargins(36, 36, 36, 36)
        card_lay.setSpacing(22)

        logo_row = QHBoxLayout()
        logo_row.setSpacing(10)
        icon_lbl = QLabel("⬡")
        icon_lbl.setStyleSheet("color: #7c3aed; font-size: 28px; background: transparent;")
        logo_row.addWidget(icon_lbl)
        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        lbl_title = QLabel("PORTARIA")
        lbl_title.setObjectName("login_title")
        lbl_sub = QLabel("SISTEMA DE CONTROLE DE ACESSO")
        lbl_sub.setObjectName("login_subtitle")
        title_col.addWidget(lbl_title)
        title_col.addWidget(lbl_sub)
        logo_row.addLayout(title_col)
        logo_row.addStretch()
        card_lay.addLayout(logo_row)
        card_lay.addWidget(make_divider())

        form_lay = QVBoxLayout()
        form_lay.setSpacing(14)

        lbl_u = QLabel("USUÁRIO")
        lbl_u.setObjectName("login_form_label")
        form_lay.addWidget(lbl_u)
        self.campo_usuario = styled_input("Digite seu usuário...")
        self.campo_usuario.setMinimumHeight(44)
        self.campo_usuario.returnPressed.connect(self._tentar_login)
        form_lay.addWidget(self.campo_usuario)

        lbl_s = QLabel("SENHA")
        lbl_s.setObjectName("login_form_label")
        form_lay.addWidget(lbl_s)
        self.campo_senha = styled_input("Digite sua senha...")
        self.campo_senha.setEchoMode(QLineEdit.EchoMode.Password)
        self.campo_senha.setMinimumHeight(44)
        self.campo_senha.returnPressed.connect(self._tentar_login)
        form_lay.addWidget(self.campo_senha)
        card_lay.addLayout(form_lay)

        self.lbl_erro = QLabel()
        self.lbl_erro.setObjectName("login_error")
        self.lbl_erro.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_erro.setWordWrap(True)
        self.lbl_erro.hide()
        card_lay.addWidget(self.lbl_erro)

        btn_entrar = QPushButton("🔓  Entrar")
        btn_entrar.setObjectName("btn_login")
        btn_entrar.setMinimumHeight(48)
        btn_entrar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_entrar.clicked.connect(self._tentar_login)
        card_lay.addWidget(btn_entrar)

        lbl_rodape = QLabel("v2.1  •  Acesso restrito")
        lbl_rodape.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_rodape.setStyleSheet("color: #2d2060; font-size: 10px; background: transparent;")
        card_lay.addWidget(lbl_rodape)

        center_lay.addWidget(card)
        center_lay.addStretch()
        outer.addWidget(center)

    def _tentar_login(self):
        usuario = self.campo_usuario.text().strip()
        senha   = self.campo_senha.text().strip()
        if usuario == LOGIN_USUARIO and senha == LOGIN_SENHA:
            self.lbl_erro.hide()
            self.login_aceito.emit()
        else:
            self.lbl_erro.setText(
                "⚠  Usuário ou senha incorretos.\nVerifique suas credenciais e tente novamente."
            )
            self.lbl_erro.show()
            self.campo_senha.clear()
            self.campo_senha.setFocus()

    def _centralizar(self):
        screen = QApplication.primaryScreen()
        if screen:
            geo = screen.availableGeometry()
            self.move(
                (geo.width()  - self.width())  // 2,
                (geo.height() - self.height()) // 2,
            )


# =============================================================================
# SIDEBAR — sem alterações
# =============================================================================

class SidebarButton(QPushButton):
    def __init__(self, icon, label, parent=None):
        super().__init__(f"  {icon}  {label}", parent)
        self.setObjectName("nav_item")
        self.setMinimumHeight(42)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

    def set_active(self, active: bool):
        self.setObjectName("nav_item_active" if active else "nav_item")
        self.style().unpolish(self)
        self.style().polish(self)


class Sidebar(QWidget):
    page_changed = pyqtSignal(int)

    NAV_ITEMS = [
        ("⬡", "Dashboard"),
        ("✦", "Cadastrar Visitante"),
        ("→", "Registrar Entrada"),
        ("◉", "Visitantes Dentro"),
        ("≡", "Histórico"),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("sidebar")
        self.setFixedWidth(230)
        self._buttons = []
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        header = QWidget()
        header.setObjectName("sidebar_header")
        header.setFixedHeight(64)
        h_lay = QVBoxLayout(header)
        h_lay.setContentsMargins(20, 14, 20, 14)
        h_lay.setSpacing(1)
        lbl_name = QLabel("⬡  PORTARIA")
        lbl_name.setObjectName("system_name")
        lbl_sub = QLabel("SISTEMA DE CONTROLE")
        lbl_sub.setObjectName("system_subtitle")
        h_lay.addWidget(lbl_name)
        h_lay.addWidget(lbl_sub)
        layout.addWidget(header)

        layout.addSpacing(16)

        nav_lbl = QLabel("  NAVEGAÇÃO")
        nav_lbl.setObjectName("section_title")
        nav_lbl.setContentsMargins(20, 0, 0, 4)
        layout.addWidget(nav_lbl)
        layout.addSpacing(4)

        nav_container = QWidget()
        nav_container.setStyleSheet("background: transparent;")
        nav_layout = QVBoxLayout(nav_container)
        nav_layout.setContentsMargins(10, 0, 10, 0)
        nav_layout.setSpacing(2)

        for idx, (icon, label) in enumerate(self.NAV_ITEMS):
            btn = SidebarButton(icon, label)
            btn.clicked.connect(lambda checked, i=idx: self._select(i))
            nav_layout.addWidget(btn)
            self._buttons.append(btn)

        layout.addWidget(nav_container)
        layout.addStretch()

        footer = QWidget()
        footer.setStyleSheet("background: transparent; border-top: 1px solid #1e1b2e;")
        f_lay = QVBoxLayout(footer)
        f_lay.setContentsMargins(20, 12, 20, 16)
        lbl_ver = QLabel("v2.1  •  Sistema Ativo")
        lbl_ver.setStyleSheet("color: #2d2060; font-size: 10px; background: transparent;")
        f_lay.addWidget(lbl_ver)
        layout.addWidget(footer)

        self._select(0)

    def _select(self, index):
        for i, btn in enumerate(self._buttons):
            btn.set_active(i == index)
        self.page_changed.emit(index)


# =============================================================================
# TOPBAR — sem alterações
# =============================================================================

class TopBar(QWidget):
    TITLES = [
        ("⬡  Dashboard",          "Visão geral do sistema"),
        ("✦  Cadastrar Visitante", "Adicione novos visitantes ao sistema"),
        ("→  Registrar Entrada",   "Registre a entrada de visitantes"),
        ("◉  Visitantes Dentro",   "Monitore quem está no local agora"),
        ("≡  Histórico",           "Consulte o histórico completo de visitas"),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("topbar")
        self.setFixedHeight(56)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(24, 0, 24, 0)
        layout.setSpacing(12)

        self.lbl_title    = QLabel()
        self.lbl_title.setObjectName("page_title")
        self.lbl_subtitle = QLabel()
        self.lbl_subtitle.setObjectName("page_subtitle")

        title_col = QVBoxLayout()
        title_col.setSpacing(1)
        title_col.addWidget(self.lbl_title)
        title_col.addWidget(self.lbl_subtitle)
        layout.addLayout(title_col)
        layout.addStretch()

        self.badge_timer  = QLabel("⟳  Auto-refresh: 10s")
        self.badge_timer.setObjectName("badge_timer")
        self.badge_online = QLabel("●  Online")
        self.badge_online.setObjectName("badge_online")
        self.lbl_clock    = QLabel()
        self.lbl_clock.setStyleSheet("color: #3a3360; font-size: 12px; background: transparent;")

        layout.addWidget(self.badge_timer)
        layout.addWidget(self.badge_online)
        layout.addWidget(self.lbl_clock)

        timer = QTimer(self)
        timer.setInterval(1000)
        timer.timeout.connect(self._update_clock)
        timer.start()
        self._update_clock()
        self.set_page(0)

    def set_page(self, index):
        title, subtitle = self.TITLES[index]
        self.lbl_title.setText(title)
        self.lbl_subtitle.setText(subtitle)

    def _update_clock(self):
        self.lbl_clock.setText(datetime.now().strftime("  %H:%M:%S"))


# =============================================================================
# PAGE 0: DASHBOARD — sem alterações
# =============================================================================

class DashboardPage(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("content_area")
        self._build()

    def _build(self):
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        container = QWidget()
        container.setObjectName("content_area")
        scroll.setWidget(container)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(20)

        lbl = QLabel("RESUMO GERAL")
        lbl.setObjectName("section_title")
        layout.addWidget(lbl)

        stats_row = QHBoxLayout()
        stats_row.setSpacing(14)
        self.stat_widgets = []

        stat_defs = [
            ("👤", "Visitantes Cadastrados", "#7c3aed"),
            ("🟢", "Dentro Agora",           "#059669"),
            ("📅", "Visitas Hoje",           "#f59e0b"),
            ("🏢", "Funcionários",           "#0284c7"),
        ]
        for icon, label, color in stat_defs:
            card = QWidget()
            card.setObjectName("stat_card")
            card.setCursor(Qt.CursorShape.PointingHandCursor)
            c_lay = QHBoxLayout(card)
            c_lay.setContentsMargins(18, 16, 18, 16)
            c_lay.setSpacing(14)

            icon_lbl = QLabel(icon)
            icon_lbl.setFixedSize(40, 40)
            icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            icon_lbl.setStyleSheet("font-size: 22px; background: transparent;")
            c_lay.addWidget(icon_lbl)

            text_col = QVBoxLayout()
            text_col.setSpacing(2)
            val_lbl = QLabel("—")
            val_lbl.setStyleSheet(
                f"color: {color}; font-size: 28px; font-weight: bold; background: transparent;"
            )
            lab_lbl = QLabel(label)
            lab_lbl.setObjectName("stat_label")
            text_col.addWidget(val_lbl)
            text_col.addWidget(lab_lbl)
            c_lay.addLayout(text_col)
            c_lay.addStretch()
            stats_row.addWidget(card)
            self.stat_widgets.append(val_lbl)

        layout.addLayout(stats_row)
        layout.addSpacing(4)

        recent_lbl = QLabel("ATIVIDADE RECENTE")
        recent_lbl.setObjectName("section_title")
        layout.addWidget(recent_lbl)

        card, card_layout = make_card(padding=0, spacing=0)
        self.recent_table = QTableWidget(0, 5)
        self.recent_table.setHorizontalHeaderLabels(
            ["Visitante", "Funcionário", "Setor", "Entrada", "Status"]
        )
        self.recent_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.recent_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.recent_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.recent_table.setAlternatingRowColors(True)
        self.recent_table.setMinimumHeight(300)
        card_layout.addWidget(self.recent_table)
        layout.addWidget(card)
        layout.addStretch()

    def refresh(self):
        try:
            stats = buscar_stats()
            for i, val in enumerate(stats):
                self.stat_widgets[i].setText(str(val))

            registros = buscar_historico_visitas()[:15]
            self.recent_table.setRowCount(0)
            for idx, (id_e, visitante, funcionario, setor, descricao, entrada, saida, porteiro) in enumerate(registros):
                self.recent_table.insertRow(idx)
                em     = saida is None
                status = "🟢 Ativo" if em else "✓ Concluído"
                entfmt = entrada.strftime("%d/%m %H:%M") if isinstance(entrada, datetime) else str(entrada)
                for col, txt in enumerate([visitante, funcionario, setor, entfmt, status]):
                    item = QTableWidgetItem(txt)
                    if em and col == 4:
                        item.setForeground(QColor("#4ade80"))
                        item.setBackground(QColor("#0a1f12"))
                    elif not em and col == 4:
                        item.setForeground(QColor("#4b4580"))
                    self.recent_table.setItem(idx, col, item)
        except Exception as e:
            print(f"[ERRO] DashboardPage.refresh: {e}")


# =============================================================================
# PAGE 1: CADASTRAR VISITANTE — sem alterações
# =============================================================================

class CadastrarPage(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("content_area")
        self._build()

    def _build(self):
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        container = QWidget()
        container.setObjectName("content_area")
        scroll.setWidget(container)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(20)

        card, card_lay = make_card(padding=28, spacing=20)
        card.setMaximumWidth(620)

        lbl = QLabel("DADOS DO VISITANTE")
        lbl.setObjectName("section_title")
        card_lay.addWidget(lbl)

        form = QFormLayout()
        form.setSpacing(14)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignLeft)
        form.setFormAlignment(Qt.AlignmentFlag.AlignLeft)

        self.campo_nome = styled_input("Nome completo do visitante...")
        form.addRow(form_label("NOME"), self.campo_nome)

        self.campo_rg = styled_input("Número do RG...")
        form.addRow(form_label("RG"), self.campo_rg)

        card_lay.addLayout(form)
        card_lay.addWidget(make_divider())

        btn_salvar = primary_button("💾  Salvar Visitante", height=44)
        btn_salvar.clicked.connect(self._salvar)
        card_lay.addWidget(btn_salvar)

        layout.addWidget(card)
        layout.addStretch()

    def _salvar(self):
        nome = self.campo_nome.text().strip()
        rg   = self.campo_rg.text().strip()
        sucesso, mensagem = tentar_cadastrar_visitante(nome, rg)
        if sucesso:
            QMessageBox.information(self, "Sucesso", mensagem)
            self.campo_nome.clear()
            self.campo_rg.clear()
        else:
            QMessageBox.warning(self, "Atenção", mensagem)

    def refresh(self):
        pass


# =============================================================================
# PAGE 2: REGISTRAR ENTRADA
# ── MODIFICADA: adicionado QLineEdit de filtro acima do combo_visitante ──────
# =============================================================================

class EntradaPage(QWidget):
    """
    Alterações em relação à versão anterior:
    - Novo QLineEdit `campo_filtro_visitante` acima do combo_visitante
    - `_visitantes_cache`: lista completa [(id, texto)] carregada do banco
    - `_filtrar_visitantes()`: filtra o cache em memória e repopula o combo
    - Nenhuma outra lógica foi alterada
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("content_area")
        self._visitantes_cache = []  # [(id, texto_exibido), ...]
        self._build()

    # ── helpers privados de filtro ────────────────────────────────────────────

    def _filtrar_visitantes(self):
        """
        Filtra self._visitantes_cache pelo texto digitado em campo_filtro_visitante.
        Opera 100% em memória — zero queries ao banco.
        Busca por substring em nome E RG (case-insensitive).
        Atualiza o contador de resultados.
        """
        termo = self.campo_filtro_visitante.text().strip().lower()

        self.combo_visitante.blockSignals(True)
        self.combo_visitante.clear()

        for id_, texto in self._visitantes_cache:
            if not termo or termo in texto.lower():
                self.combo_visitante.addItem(texto, userData=id_)

        self.combo_visitante.blockSignals(False)

        total    = len(self._visitantes_cache)
        exibidos = self.combo_visitante.count()
        if termo:
            self.lbl_filtro_contador.setText(
                f"{exibidos} de {total} visitante{'s' if total != 1 else ''}"
            )
        else:
            self.lbl_filtro_contador.setText(
                f"{total} visitante{'s' if total != 1 else ''} cadastrado{'s' if total != 1 else ''}"
            )

    # ── build ─────────────────────────────────────────────────────────────────

    def _build(self):
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        container = QWidget()
        container.setObjectName("content_area")
        scroll.setWidget(container)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(20)

        card, card_lay = make_card(padding=28, spacing=20)
        card.setMaximumWidth(620)

        lbl = QLabel("DADOS DA ENTRADA")
        lbl.setObjectName("section_title")
        card_lay.addWidget(lbl)

        form = QFormLayout()
        form.setSpacing(16)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignLeft)

        # ── Campo de filtro (NOVO) ──────────────────────────────────────────
        filtro_col = QVBoxLayout()
        filtro_col.setSpacing(4)

        self.campo_filtro_visitante = styled_input("🔍  Buscar por nome ou RG...")
        self.campo_filtro_visitante.setMinimumHeight(36)
        self.campo_filtro_visitante.textChanged.connect(self._filtrar_visitantes)
        filtro_col.addWidget(self.campo_filtro_visitante)

        self.lbl_filtro_contador = QLabel("")
        self.lbl_filtro_contador.setObjectName("filtro_contador")
        filtro_col.addWidget(self.lbl_filtro_contador)

        self.combo_visitante = styled_combo()
        filtro_col.addWidget(self.combo_visitante)

        form.addRow(form_label("VISITANTE"), filtro_col)
        # ── Fim campo de filtro ────────────────────────────────────────────

        self.combo_funcionario = styled_combo()
        form.addRow(form_label("FUNCIONÁRIO RESPONSÁVEL"), self.combo_funcionario)

        self.campo_porteiro = styled_input("Nome do porteiro responsável pelo acesso...")
        form.addRow(form_label("PORTEIRO"), self.campo_porteiro)

        self.campo_descricao = styled_input(
            "Objetivo da visita (ex: manutenção, reunião, visita técnica...)"
        )
        form.addRow(form_label("DESCRIÇÃO"), self.campo_descricao)

        card_lay.addLayout(form)

        btn_att = secondary_button("🔄  Atualizar Listas", height=36)
        btn_att.clicked.connect(self._carregar_combos)
        card_lay.addWidget(btn_att)

        card_lay.addWidget(make_divider())

        btn_entrada = success_button("✅  Registrar Entrada", height=44)
        btn_entrada.clicked.connect(self._confirmar)
        card_lay.addWidget(btn_entrada)

        layout.addWidget(card)
        layout.addStretch()
        self._carregar_combos()

    # ── carregar dados ────────────────────────────────────────────────────────

    def _carregar_combos(self):
        """
        Carrega visitantes e funcionários do banco UMA única vez.
        Visitantes são guardados em _visitantes_cache para filtragem em memória.
        O campo de filtro é limpo para exibir a lista completa.
        """
        try:
            # Visitantes: preenche cache e limpa o filtro
            self._visitantes_cache = [
                (id_, f"{nome}  (RG: {rg})")
                for id_, nome, rg in buscar_visitantes()
            ]
            self.campo_filtro_visitante.blockSignals(True)
            self.campo_filtro_visitante.clear()
            self.campo_filtro_visitante.blockSignals(False)
            self._filtrar_visitantes()  # popula combo a partir do cache

            # Funcionários: lógica original, sem alteração
            self.combo_funcionario.clear()
            for id_, nome, setor in buscar_funcionarios():
                self.combo_funcionario.addItem(f"{nome}  —  {setor}", userData=id_)

        except Exception as e:
            print(f"[ERRO] EntradaPage._carregar_combos: {e}")
            QMessageBox.critical(self, "Erro", "Erro ao carregar listas.")

    # ── confirmar entrada ─────────────────────────────────────────────────────

    def _confirmar(self):
        """Lógica original preservada. Usa findText para garantir ID correto."""
        texto = self.combo_visitante.currentText()
        indice = self.combo_visitante.findText(texto)

        if indice == -1:
            QMessageBox.warning(self, "Atenção", "Selecione um visitante válido da lista!")
            return

        v_id = self.combo_visitante.itemData(indice)
        f_id = self.combo_funcionario.currentData()

        sucesso, mensagem = tentar_registrar_entrada(
            v_id,
            f_id,
            self.campo_descricao.text(),
            self.campo_porteiro.text(),
        )

        if sucesso:
            horario = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            QMessageBox.information(self, "Sucesso", f"{mensagem}\nHorário: {horario}")
            self.campo_descricao.clear()
            self.campo_porteiro.clear()
            self.campo_filtro_visitante.clear()  # limpa filtro após sucesso
        else:
            QMessageBox.warning(self, "Atenção", mensagem)

    def refresh(self):
        pass


# =============================================================================
# PAGE 3: VISITANTES DENTRO
# ── MODIFICADA: botões de exportação CSV / Excel na toolbar ──────────────────
# =============================================================================

class DentroPage(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("content_area")
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(10)
        lbl = QLabel("VISITANTES NO LOCAL")
        lbl.setObjectName("section_title")
        toolbar.addWidget(lbl)
        toolbar.addStretch()

        btn_att = secondary_button("🔄  Atualizar", height=34)
        btn_att.clicked.connect(self.refresh)
        toolbar.addWidget(btn_att)

        # ── Exportação (NOVO) ──────────────────────────────────────────────
        btn_csv = secondary_button("⬇  CSV", height=34)
        btn_csv.setMaximumWidth(90)
        btn_csv.setToolTip("Exportar lista atual para CSV")
        btn_csv.clicked.connect(
            lambda: exportar_tabela_csv(self, self.tabela, "visitantes_dentro")
        )
        toolbar.addWidget(btn_csv)

        btn_xlsx = secondary_button("⬇  Excel", height=34)
        btn_xlsx.setMaximumWidth(90)
        btn_xlsx.setToolTip("Exportar lista atual para Excel (requer openpyxl)")
        btn_xlsx.clicked.connect(
            lambda: exportar_tabela_excel(self, self.tabela, "visitantes_dentro")
        )
        toolbar.addWidget(btn_xlsx)
        # ── Fim exportação ─────────────────────────────────────────────────

        btn_saida = danger_button("🚪  Registrar Saída", height=34)
        btn_saida.clicked.connect(self._confirmar_saida)
        toolbar.addWidget(btn_saida)
        layout.addLayout(toolbar)

        self.counter_lbl = QLabel("— visitantes dentro")
        self.counter_lbl.setStyleSheet("color: #4b4580; font-size: 12px; background: transparent;")
        layout.addWidget(self.counter_lbl)

        self.tabela = QTableWidget(0, 5)
        self.tabela.setHorizontalHeaderLabels(["ID", "Visitante", "Funcionário", "Setor", "Entrada"])
        self.tabela.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabela.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabela.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabela.setAlternatingRowColors(True)
        layout.addWidget(self.tabela)

        self.refresh()

    def refresh(self):
        try:
            registros = buscar_visitantes_dentro()
            self.tabela.setRowCount(0)
            for idx, (id_e, visitante, funcionario, setor, entrada) in enumerate(registros):
                self.tabela.insertRow(idx)
                entfmt = entrada.strftime("%d/%m/%Y %H:%M") if isinstance(entrada, datetime) else str(entrada)
                for col, txt in enumerate([str(id_e), visitante, funcionario, setor, entfmt]):
                    item = QTableWidgetItem(txt)
                    if col == 1:
                        item.setForeground(QColor("#c4b5fd"))
                    self.tabela.setItem(idx, col, item)

            n = len(registros)
            self.counter_lbl.setText(f"{n} visitante{'s' if n != 1 else ''} dentro agora")
        except Exception as e:
            print(f"[ERRO] DentroPage.refresh: {e}")

    def _confirmar_saida(self):
        linha = self.tabela.currentRow()
        if linha == -1:
            QMessageBox.warning(self, "Atenção", "Selecione um visitante na tabela!")
            return

        id_entrada = int(self.tabela.item(linha, 0).text())
        nome       = self.tabela.item(linha, 1).text()

        resp = QMessageBox.question(
            self, "Confirmar Saída", f"Registrar saída de '{nome}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if resp == QMessageBox.StandardButton.Yes:
            try:
                registrar_saida(id_entrada)
                QMessageBox.information(self, "Sucesso", f"Saída de '{nome}' registrada!")
                self.refresh()
            except Exception as e:
                print(f"[ERRO] DentroPage._confirmar_saida: {e}")
                QMessageBox.critical(self, "Erro", "Erro ao registrar saída.")


# =============================================================================
# PAGE 4: HISTÓRICO DE VISITAS
# ── MODIFICADA: botões de exportação CSV / Excel na toolbar ──────────────────
# =============================================================================

class HistoricoPage(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("content_area")
        self._cache = []
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(10)
        lbl = QLabel("HISTÓRICO COMPLETO")
        lbl.setObjectName("section_title")
        toolbar.addWidget(lbl)
        toolbar.addStretch()

        btn_att = secondary_button("🔄  Atualizar", height=34)
        btn_att.clicked.connect(self.refresh)
        toolbar.addWidget(btn_att)

        # ── Exportação (NOVO) ──────────────────────────────────────────────
        btn_csv = secondary_button("⬇  CSV", height=34)
        btn_csv.setMaximumWidth(90)
        btn_csv.setToolTip("Exportar registros filtrados para CSV")
        btn_csv.clicked.connect(
            lambda: exportar_tabela_csv(self, self.tabela, "historico_visitas")
        )
        toolbar.addWidget(btn_csv)

        btn_xlsx = secondary_button("⬇  Excel", height=34)
        btn_xlsx.setMaximumWidth(90)
        btn_xlsx.setToolTip("Exportar registros filtrados para Excel (requer openpyxl)")
        btn_xlsx.clicked.connect(
            lambda: exportar_tabela_excel(self, self.tabela, "historico_visitas")
        )
        toolbar.addWidget(btn_xlsx)
        # ── Fim exportação ─────────────────────────────────────────────────

        layout.addLayout(toolbar)

        hint_lbl = QLabel("💡  Clique duplo em qualquer linha para ver todos os detalhes da visita")
        hint_lbl.setStyleSheet(
            "color: #4b4580; font-size: 11px; background: #13101f; "
            "border: 1px solid #1e1b2e; border-radius: 6px; padding: 6px 12px;"
        )
        layout.addWidget(hint_lbl)

        search_bar = QHBoxLayout()
        search_bar.setSpacing(10)

        self.campo_busca = styled_input("🔍  Filtrar por nome do visitante...")
        self.campo_busca.setMaximumHeight(38)
        self.campo_busca.textChanged.connect(self._filtrar)
        search_bar.addWidget(self.campo_busca)

        self.campo_data = styled_input("📅  dd/mm/aaaa")
        self.campo_data.setMaximumWidth(140)
        self.campo_data.setMaximumHeight(38)
        self.campo_data.textChanged.connect(self._filtrar)
        search_bar.addWidget(self.campo_data)

        btn_clear = secondary_button("✖  Limpar", height=38)
        btn_clear.setMaximumWidth(100)
        btn_clear.clicked.connect(lambda: (self.campo_busca.clear(), self.campo_data.clear()))
        search_bar.addWidget(btn_clear)
        layout.addLayout(search_bar)

        self.tabela = QTableWidget(0, 8)
        self.tabela.setHorizontalHeaderLabels(
            ["ID", "Visitante", "Funcionário", "Porteiro", "Setor", "Descrição", "Entrada", "Saída"]
        )
        self.tabela.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabela.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabela.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabela.setAlternatingRowColors(True)
        self.tabela.cellDoubleClicked.connect(self._abrir_detalhes)
        layout.addWidget(self.tabela)

        self.refresh()

    def refresh(self):
        try:
            self._cache = buscar_historico_visitas()
            self._filtrar()
        except Exception as e:
            print(f"[ERRO] HistoricoPage.refresh: {e}")

    def _filtrar(self):
        tn = self.campo_busca.text().strip().lower()
        td = self.campo_data.text().strip()
        filtrados = []
        for r in self._cache:
            _, visitante, _, _, _, entrada, _, _ = r
            if tn and tn not in visitante.lower():
                continue
            if td:
                ef = entrada.strftime("%d/%m/%Y") if isinstance(entrada, datetime) else str(entrada)
                if td not in ef:
                    continue
            filtrados.append(r)
        self._popular_tabela(filtrados)

    def _popular_tabela(self, registros):
        self.tabela.setRowCount(0)
        for idx, (id_e, visitante, funcionario, setor, descricao, entrada, saida, porteiro) in enumerate(registros):
            self.tabela.insertRow(idx)
            em      = saida is None
            entfmt  = entrada.strftime("%d/%m/%Y %H:%M") if isinstance(entrada, datetime) else str(entrada)
            saidfmt = (
                "🟢 Em andamento" if em
                else (saida.strftime("%d/%m/%Y %H:%M") if isinstance(saida, datetime) else str(saida))
            )
            desc = descricao or "—"
            port = porteiro  or "—"

            for col, txt in enumerate([str(id_e), visitante, funcionario, port, setor, desc, entfmt, saidfmt]):
                item = QTableWidgetItem(txt)
                if em:
                    item.setBackground(QColor("#0a1f12"))
                    item.setForeground(QColor("#4ade80"))
                elif col == 7:
                    item.setForeground(QColor("#4b4580"))
                self.tabela.setItem(idx, col, item)

    def _abrir_detalhes(self, row: int, _col: int):
        if row < 0 or row >= self.tabela.rowCount():
            return

        id_real  = int(self.tabela.item(row, 0).text())
        registro = next((r for r in self._cache if r[0] == id_real), None)
        if registro is None:
            return

        id_e, visitante, funcionario, setor, descricao, entrada, saida, porteiro = registro

        em      = saida is None
        entfmt  = entrada.strftime("%d/%m/%Y %H:%M:%S") if isinstance(entrada, datetime) else str(entrada)
        saidfmt = None if em else (
            saida.strftime("%d/%m/%Y %H:%M:%S") if isinstance(saida, datetime) else str(saida)
        )
        tempo = formatar_tempo_permanencia(entrada, None if em else saida)

        dados = {
            "id":          id_e,
            "visitante":   visitante,
            "funcionario": funcionario,
            "porteiro":    porteiro or "",
            "setor":       setor,
            "descricao":   descricao or "",
            "entrada":     entfmt,
            "saida":       saidfmt,
            "permanencia": tempo,
        }

        dlg = DetalhesVisitaDialog(dados, parent=self)
        dlg.exec()


# =============================================================================
# JANELA PRINCIPAL — sem alterações
# =============================================================================

class JanelaPortaria(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Sistema de Portaria")
        self.setStyleSheet(STYLESHEET)
        self._build()
        self._setup_timer()
        self.showMaximized()

    def _build(self):
        root = QWidget()
        self.setCentralWidget(root)
        root_lay = QHBoxLayout(root)
        root_lay.setContentsMargins(0, 0, 0, 0)
        root_lay.setSpacing(0)

        self.sidebar = Sidebar()
        self.sidebar.page_changed.connect(self._switch_page)
        root_lay.addWidget(self.sidebar)

        main_col = QWidget()
        main_col.setObjectName("main_area")
        main_lay = QVBoxLayout(main_col)
        main_lay.setContentsMargins(0, 0, 0, 0)
        main_lay.setSpacing(0)

        self.topbar = TopBar()
        main_lay.addWidget(self.topbar)

        self.stack = QStackedWidget()
        self.stack.setObjectName("content_area")

        self.pages = [
            DashboardPage(),
            CadastrarPage(),
            EntradaPage(),
            DentroPage(),
            HistoricoPage(),
        ]
        for page in self.pages:
            self.stack.addWidget(page)

        main_lay.addWidget(self.stack, 1)
        root_lay.addWidget(main_col, 1)

        self._switch_page(0)

    def _switch_page(self, index):
        self.stack.setCurrentIndex(index)
        self.topbar.set_page(index)
        self.pages[index].refresh()

    def _setup_timer(self):
        self.timer = QTimer(self)
        self.timer.setInterval(10_000)
        self.timer.timeout.connect(
            lambda: self.pages[self.stack.currentIndex()].refresh()
        )
        self.timer.start()